"""Export service — generate PDF, Excel, and CSV reports."""

import csv
import io
from datetime import date, datetime
from typing import Optional, List, Dict, Any
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session


def _get_dashboard_data(db: Session, date_from: Optional[date], date_to: Optional[date]) -> Dict[str, Any]:
    """Fetch dashboard data for export."""
    from app.services.dashboard_service import get_kpi_summary, get_payment_breakdown
    kpi = get_kpi_summary(db, date_from, date_to)
    payments = get_payment_breakdown(db, date_from, date_to)
    return {
        "kpi": kpi,
        "payments": payments,
        "date_from": str(date_from) if date_from else "N/A",
        "date_to": str(date_to) if date_to else "N/A",
    }


def generate_pdf(
    db: Session,
    report_type: str = "dashboard",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> io.BytesIO:
    """Generate a PDF report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=20, textColor=colors.HexColor("#1e293b"),
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle', parent=styles['Normal'],
        fontSize=12, textColor=colors.HexColor("#64748b"),
        spaceAfter=10,
    )

    elements = []

    # Title
    elements.append(Paragraph("BI Dashboard Report", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Period: {date_from or 'Today'} to {date_to or 'Today'}",
        subtitle_style
    ))
    elements.append(Spacer(1, 20))

    data = _get_dashboard_data(db, date_from, date_to)
    kpi = data["kpi"]

    # KPI Table
    kpi_data = [
        ["Metric", "Value"],
        ["Total Sales", f"KES {kpi.total_sales:,.2f}"],
        ["Total Purchases", f"KES {kpi.total_purchases:,.2f}"],
        ["Gross Profit", f"KES {kpi.gross_profit:,.2f}"],
        ["Net Profit", f"KES {kpi.net_profit:,.2f}"],
        ["Cash in Hand", f"KES {kpi.cash_in_hand:,.2f}"],
        ["Total Customers", f"{kpi.customer_count:,}"],
        ["Total Transactions", f"{kpi.transaction_count:,}"],
        ["Avg Transaction", f"KES {kpi.avg_transaction_value:,.2f}"],
    ]

    table = Table(kpi_data, colWidths=[3*inch, 3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6366f1")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    # Payment Breakdown Table
    if data["payments"]:
        elements.append(Paragraph("Payment Method Breakdown", styles['Heading2']))
        pay_data = [["Payment Method", "Amount (KES)", "Percentage"]]
        for p in data["payments"]:
            pay_data.append([p.method, f"{p.amount:,.2f}", f"{p.percentage}%"])

        pay_table = Table(pay_data, colWidths=[3*inch, 2*inch, 1.5*inch])
        pay_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(pay_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel(
    db: Session,
    report_type: str = "dashboard",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> io.BytesIO:
    """Generate an Excel report with formatted sheets."""
    wb = Workbook()

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    data = _get_dashboard_data(db, date_from, date_to)
    kpi = data["kpi"]

    # KPI Sheet
    ws = wb.active
    ws.title = "Dashboard KPIs"
    ws.append(["BI Dashboard Report"])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=16, color="1E293B")
    ws.append([f"Period: {data['date_from']} to {data['date_to']}"])
    ws.append([])

    headers = ["Metric", "Value"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    kpi_rows = [
        ("Total Sales", kpi.total_sales),
        ("Total Purchases", kpi.total_purchases),
        ("Total Expenses", kpi.total_expenses),
        ("Gross Profit", kpi.gross_profit),
        ("Net Profit", kpi.net_profit),
        ("Cash in Hand", kpi.cash_in_hand),
        ("Total Customers", kpi.customer_count),
        ("Total Transactions", kpi.transaction_count),
        ("Avg Transaction Value", kpi.avg_transaction_value),
    ]
    for metric, value in kpi_rows:
        ws.append([metric, value])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    # Payment Breakdown Sheet
    ws2 = wb.create_sheet("Payment Breakdown")
    ws2.append(["Payment Method", "Amount (KES)", "Percentage (%)"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for p in data["payments"]:
        ws2.append([p.method, p.amount, p.percentage])

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_csv(
    db: Session,
    report_type: str = "dashboard",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> io.StringIO:
    """Generate a CSV report."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    data = _get_dashboard_data(db, date_from, date_to)
    kpi = data["kpi"]

    writer.writerow(["BI Dashboard Report"])
    writer.writerow([f"Period: {data['date_from']} to {data['date_to']}"])
    writer.writerow([])

    writer.writerow(["Metric", "Value"])
    writer.writerow(["Total Sales", kpi.total_sales])
    writer.writerow(["Total Purchases", kpi.total_purchases])
    writer.writerow(["Total Expenses", kpi.total_expenses])
    writer.writerow(["Gross Profit", kpi.gross_profit])
    writer.writerow(["Net Profit", kpi.net_profit])
    writer.writerow(["Cash in Hand", kpi.cash_in_hand])
    writer.writerow(["Total Customers", kpi.customer_count])
    writer.writerow(["Total Transactions", kpi.transaction_count])
    writer.writerow([])

    writer.writerow(["Payment Method", "Amount", "Percentage"])
    for p in data["payments"]:
        writer.writerow([p.method, p.amount, p.percentage])

    buffer.seek(0)
    return buffer
